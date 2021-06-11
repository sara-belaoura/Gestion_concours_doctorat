from django import http
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.http import HttpResponse  
from app.functions.functions import handle_uploaded_file  
from .forms import StudentForm   
import openpyxl
from django.http import HttpResponse
from django.contrib.auth.forms import UserCreationForm
from .models import InputForm
from django.contrib.auth import authenticate,login
from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from django.template import RequestContext

def home_page(request):
    return render(request,'app/home.html',{})

def signup(request):
    if request.method == 'POST':
        form = InputForm(request.POST)
        if form.is_valid():
            form.save()
            first_name = form.cleaned_data.get('first_name')
            last_name = form.cleaned_data.get('last_name')
            email = form.cleaned_data.get('email')
            username = form.cleaned_data.get('username')
            raw_password = form.cleaned_data.get('password1')
            user = authenticate(username=username,password=raw_password)
            login(request,user)
           
            return redirect('home_page')
    else:
        form = InputForm()
    return render(request, 'app/signup.html', {'form': form})
def user_login(request):
    context = RequestContext(request)
    if request.method == 'POST':
          email = request.POST['email']
          password = request.POST['password']
          user = authenticate(email=email, password=password)
          if user is not None:
              
                  login(request,user)
                  # Redirect to index page.
                  return redirect('home_page')
              
                  
          else:
              # Return an 'invalid login' error message.
              print ( "invalid login details " + email + " " + password)
              #return render(request,'app/login.html', {}, context)
     

def index(request) :
    return render(request,'index.html', {})

def recupererList(request) :
    return render(request,'de/recupererList.html', {})

def affecterSalle(request) :   
    if "GET" == request.method:
        return render(request,'de/affecterSalles.html', {})   
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]

        

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        nb_total=0
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
            nb_total=nb_total+1
    
    request.session['excel'] = excel_data
            
    return render(request, "de/affecterSalles.html", {"excel_data":excel_data,"excel_file":excel_file,"nb_total":nb_total})

def affecterSurveillant(request) :
    if "GET" == request.method:
        return render(request,'de/AffectationSurveillant.html', {}) 
    else: 
      return render(request,'de/AffectationSurveillant.html', {}) 



def validerSalles(request):
    validated=True
    excel_data=request.session['excel']
    nb_total=0
    for row in excel_data :
        nb_total=nb_total+1 
        """ each row is a row in the database"""
    
    return render(request,'de/affecterSalles.html', {"validated":validated,"excel_data":excel_data,"nb_total":nb_total})
 

"""============================================================================="""

def importFile(request): 
    
    if "GET" == request.method:
        return render(request, 'importFile.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        
        return render(request, "importFile.html", {"excel_data":excel_data})


"""---------------------------------------------------------------"""
 # Import mimetypes module
import mimetypes
# import os module
import os
# Import HttpResponse module
from django.http.response import HttpResponse


def uploadFile(request):
  
    if request.method == 'POST':  
        if request.FILES.get('file_name',False) :
            handle_uploaded_file(request.FILES['file_name'])  
            return HttpResponse("File uploaded successfuly")        
        else:  
            return HttpResponse("Error : please enter a valid file !")      
    else:  
        return render(request,"uploadFile.html")


def downloadFile(request):
    if request.method == 'POST':  
        # Define Django project base directory
        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        # Define text file name
        x=dict(request.POST.items())
        if 'SIQ' in request.POST :
            filename="liste_siq.xlsx"
        elif 'SIT' in request.POST  :
            filename="liste_sit.xlsx"
       
        # Define the full file path
        filepath = BASE_DIR + '/app/static/upload/' + filename
        # Open the file for reading content
        path = open(filepath, 'r' ,encoding="cp437")
        # Set the mime type
        mime_type, _ = mimetypes.guess_type(filepath)
        # Set the return value of the HttpResponse
        response = HttpResponse(path, content_type=mime_type)
        # Set the HTTP header for sending to browser
        response['Content-Disposition'] = "attachment; filename=%s" % filename
        # Return the response value
        return response 
          
    else: 
         return render(request,'de/recupererList.html', {})

